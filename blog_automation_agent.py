#!/usr/bin/env python3
"""
FaMED Blog Automation Agent
- Posts one blog per day
- Updates Excel tracker with tags and status
- Finds new outlier topics from FSP blogs
- Pushes changes to GitHub
"""

import os
import sys
import json
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import glob
import re
import subprocess

# Import notifications (Telegram/Email - no Spanish psychologists!)
try:
    from notification_system import UniversalNotifier
    NOTIF_ENABLED = True
except ImportError:
    NOTIF_ENABLED = False
    print("⚠️  Notifications not available (optional)")

# Configuration
BLOG_POSTS_DIR = "./blog/posts"
DRAFTS_DIR = "./blog/drafts"
EXCEL_TRACKER = "./FSP_Blog_Competitive_Analysis.xlsx"
GIT_REPO = "https://github.com/dryanez/Famedblog"

class BlogAutomationAgent:
    def __init__(self):
        self.today = datetime.now().strftime("%Y-%m-%d")
        self.tracker_df = None
        self.load_tracker()
        
        # Initialize notifier (Telegram or Email)
        if NOTIF_ENABLED:
            self.notifier = UniversalNotifier()
        else:
            self.notifier = None
    
    def load_tracker(self):
        """Load Excel tracker"""
        try:
            self.tracker_df = pd.read_excel(EXCEL_TRACKER, sheet_name='Top FSP Blog Posts')
            print(f"✅ Loaded tracker with {len(self.tracker_df)} posts")
        except Exception as e:
            print(f"❌ Error loading tracker: {e}")
            sys.exit(1)
    
    def get_next_post_to_publish(self):
        """Find the next draft post that should be published today"""
        draft_files = glob.glob(f"{BLOG_POSTS_DIR}/*.md")
        
        for file_path in sorted(draft_files):
            metadata = self.extract_metadata(file_path)
            
            # Check if this post should be published today or in the past (catch-up)
            if metadata.get('status') in ['draft', 'scheduled'] and metadata.get('date') <= self.today:
                return file_path, metadata
        
        print("ℹ️  No posts scheduled for today")
        return None, None
    
    def extract_metadata(self, file_path):
        """Extract frontmatter metadata from markdown file"""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Extract YAML frontmatter
        match = re.search(r'^---\n(.*?)\n---', content, re.DOTALL)
        if not match:
            return {}
        
        frontmatter = match.group(1)
        metadata = {}
        
        for line in frontmatter.split('\n'):
            if ': ' in line:
                key, value = line.split(': ', 1)
                key = key.strip()
                value = value.strip().strip('"\'')
                
                # Handle tags as list
                if key == 'tags':
                    metadata[key] = [t.strip().strip('"\'') for t in value.strip('[]').split(',')]
                else:
                    metadata[key] = value
        
        return metadata
    
    def update_post_status(self, file_path, new_status='published'):
        """Update post status in markdown file"""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Update status in frontmatter
        content = re.sub(
            r'status:\s*"?draft"?',
            f'status: "{new_status}"',
            content
        )
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        print(f"✅ Updated {os.path.basename(file_path)} status to: {new_status}")
    
    def update_excel_tracker(self, post_title, tags, status='Published'):
        """Update Excel tracker with post information"""
        try:
            wb = openpyxl.load_workbook(EXCEL_TRACKER)
            ws = wb['Top FSP Blog Posts']
            
            # Find matching row by title
            for row in range(2, ws.max_row + 1):
                recommended_title = ws.cell(row, 13).value  # Column M: Recommended FaMED Title
                
                if recommended_title and post_title in recommended_title:
                    # Update tags (new column)
                    ws.cell(row, 14, ', '.join(tags))  # Column N: Tags
                    ws.cell(row, 14).font = Font(size=10)
                    
                    # Update status (new column)
                    ws.cell(row, 15, status)  # Column O: Status
                    ws.cell(row, 15).fill = PatternFill(start_color='90EE90', end_color ='90EE90', fill_type='solid')
                    ws.cell(row, 15).font = Font(bold=True)
                    
                    # Update date published (new column)
                    ws.cell(row, 16, self.today)  # Column P: Date Published
                    
                    print(f"✅ Updated tracker for: {post_title}")
                    break
            
            wb.save(EXCEL_TRACKER)
            
        except Exception as e:
            print(f"⚠️  Error updating Excel: {e}")
    
    def git_push(self, commit_message):
        """Commit and push changes to GitHub"""
        try:
            # Add all changes
            subprocess.run(['git', 'add', '.'], check=True)
            
            # Commit
            subprocess.run(['git', 'commit', '-m', commit_message], check=True)
            
            # Push
            subprocess.run(['git', 'push', 'origin', 'main'], check=True)
            
            print(f"✅ Pushed to GitHub: {commit_message}")
            return True
            
        except subprocess.CalledProcessError as e:
            print(f"❌ Git error: {e}")
            return False
    
    def search_for_new_outliers(self):
        """Search for new high-performing FSP blog topics to add to queue"""
        print("\n🔍 Searching for new outlier FSP topics...")
        
        # This would typically use web scraping or API calls
        # For now, we'll create a placeholder
        
        new_topics = [
            {
                'title': 'FSP Vokabelliste: 500 Essential Medical Terms',
                'category': 'Vocabulary',
                'estimated_popularity': 85,
                'famed_adaptation': 'HIGH',
                'recommended_title': 'FaMED Essential Vocabulary: 500 Must-Know German Medical Terms'
            },
            {
                'title': 'FSP Prüfung Bayern: Regional Differences Explained',
                'category': 'Regional Guide',
                'estimated_popularity': 72,
                'famed_adaptation': 'MEDIUM',
                'recommended_title': 'FaMED State-by-State Guide: Regional Requirements in Germany'
            }
        ]
        
        return new_topics
    
    def add_outliers_to_tracker(self, new_topics):
        """Add new outlier topics to Excel tracker"""
        try:
            df_new = pd.DataFrame(new_topics)
            
            # Append to existing tracker
            wb = openpyxl.load_workbook(EXCEL_TRACKER)
            ws = wb['Top FSP Blog Posts']
            
            next_row = ws.max_row + 1
            
            for topic in new_topics:
                ws.cell(next_row, 1, next_row - 1)  # Rank
                ws.cell(next_row, 2, topic['title'])
                ws.cell(next_row, 4, topic['category'])
                ws.cell(next_row, 6, topic['estimated_popularity'])
                ws.cell(next_row, 12, topic['famed_adaptation'])
                ws.cell(next_row, 13, topic['recommended_title'])
                ws.cell(next_row, 15, 'Queued')  # Status
                
                next_row += 1
            
            wb.save(EXCEL_TRACKER)
            print(f"✅ Added {len(new_topics)} new topics to tracker")
            
        except Exception as e:
            print(f"⚠️  Error adding outliers: {e}")
    
    def run_daily_automation(self):
        """Main automation workflow"""
        print("=" * 60)
        print(f"🤖 FaMED Blog Automation Agent - {self.today}")
        print("=" * 60)
        
        # Step 1: Check if there's a post to publish today
        post_file, metadata = self.get_next_post_to_publish()
        
        if post_file:
            print(f"\n📝 Publishing: {metadata.get('title')}")
            
            # Step 2: Update post status to published
            self.update_post_status(post_file, 'published')
            
            # Step 3: Update Excel tracker
            self.update_excel_tracker(
                metadata.get('title'),
                metadata.get('tags', []),
                'Published'
            )
            
            # Step 4: Git commit and push
            commit_msg = f"Published: {metadata.get('title')} - {self.today}"
            self.git_push(commit_msg)
            
            print(f"\n✅ Successfully published blog post!")
            
            # Step 5: Send notification
            if self.notifier:
                self.notifier.notify_blog_published(
                    metadata.get('title'),
                    self.today
                )
        
        # Step 5: Search for new outlier topics (weekly - only on Mondays)
        if datetime.now().weekday() == 0:  # Monday
            new_topics = self.search_for_new_outliers()
            
            if new_topics:
                print(f"\n🎯 Found {len(new_topics)} new outlier topics")
                self.add_outliers_to_tracker(new_topics)
                
                # Commit tracker updates
                self.git_push(f"Added {len(new_topics)} new blog topic ideas - {self.today}")
                
                # Send notification
                if self.notifier:
                    self.notifier.notify_new_topics_found(len(new_topics))
        
        print("\n" + "=" * 60)
        print("✅ Automation complete!")
        print("=" * 60)

def main():
    agent = BlogAutomationAgent()
    agent.run_daily_automation()

if __name__ == "__main__":
    main()
